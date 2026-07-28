"""Export helpers — PDF, Excel, CSV."""
from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from config.settings import settings
from utils.formatters import format_fc


def export_tickets_csv(tickets, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(
            [
                "Numero",
                "Date",
                "Passager",
                "Telephone",
                "Trajet",
                "Bus",
                "Siege",
                "Prix",
                "Caissier",
                "Statut",
            ]
        )
        for t in tickets:
            w.writerow(
                [
                    t.numero,
                    t.date_vente.isoformat(),
                    t.passenger_name,
                    t.phone,
                    t.route.short_label if t.route else "",
                    t.bus.code if t.bus else "",
                    t.seat_number,
                    float(t.price),
                    t.cashier.full_name if t.cashier else "",
                    t.statut,
                ]
            )
    return path


def export_tickets_excel(tickets, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"
    ws.append(
        [
            "Numero",
            "Date",
            "Passager",
            "Telephone",
            "Trajet",
            "Bus",
            "Siege",
            "Prix",
            "Caissier",
            "Statut",
        ]
    )
    for t in tickets:
        ws.append(
            [
                t.numero,
                t.date_vente.isoformat(),
                t.passenger_name,
                t.phone,
                t.route.short_label if t.route else "",
                t.bus.code if t.bus else "",
                t.seat_number,
                float(t.price),
                t.cashier.full_name if t.cashier else "",
                t.statut,
            ]
        )
    wb.save(path)
    return path


def export_tickets_pdf(tickets, path: Path, title: str = "Historique des ventes") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    y = height - 40
    if settings.logo_path.exists():
        try:
            c.drawImage(str(settings.logo_path), 40, y - 30, 45, 35, mask="auto")
            c.setFont("Helvetica-Bold", 14)
            c.drawString(95, y, title)
            c.setFont("Helvetica", 9)
            c.drawString(95, y - 16, f"Exporté le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            y -= 45
        except Exception:
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40, y, title)
            y -= 20
            c.setFont("Helvetica", 9)
            c.drawString(40, y, f"Exporté le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            y -= 24
    else:
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, title)
        y -= 20
        c.setFont("Helvetica", 9)
        c.drawString(40, y, f"Exporté le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 24
    headers = "Numéro | Date | Passager | Tél | Trajet | Siège | Prix | Statut"
    c.setFont("Helvetica-Bold", 8)
    c.drawString(40, y, headers)
    y -= 14
    c.setFont("Helvetica", 8)
    for t in tickets:
        line = (
            f"{t.numero} | {t.date_vente} | {t.passenger_name[:20]} | {t.phone} | "
            f"{(t.route.short_label if t.route else '')[:22]} | {t.seat_number} | "
            f"{float(t.price):.0f} | {t.statut}"
        )
        if y < 40:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica", 8)
        c.drawString(40, y, line[:120])
        y -= 12
    c.save()
    return path


def export_luggage_csv(items, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(
            [
                "Numero",
                "Expediteur",
                "Destinataire",
                "Description",
                "Poids",
                "Total",
                "Bus",
                "Trajet",
                "Statut",
                "Date",
            ]
        )
        for i in items:
            w.writerow(
                [
                    i.numero,
                    i.sender_name,
                    i.recipient_name,
                    i.description,
                    float(i.poids),
                    float(i.total),
                    i.bus.code if i.bus else "",
                    i.route.short_label if i.route else "",
                    i.statut,
                    i.created_at.isoformat(sep=" ", timespec="minutes"),
                ]
            )
    return path


def export_luggage_excel(items, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bagages"
    ws.append(
        [
            "Numero",
            "Expediteur",
            "Destinataire",
            "Description",
            "Poids",
            "Total",
            "Bus",
            "Trajet",
            "Statut",
            "Date",
        ]
    )
    for i in items:
        ws.append(
            [
                i.numero,
                i.sender_name,
                i.recipient_name,
                i.description,
                float(i.poids),
                float(i.total),
                i.bus.code if i.bus else "",
                i.route.short_label if i.route else "",
                i.statut,
                i.created_at.isoformat(sep=" ", timespec="minutes"),
            ]
        )
    wb.save(path)
    return path


def export_luggage_pdf(items, path: Path, title: str = "Manifeste bagages") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    y = height - 40
    if settings.logo_path.exists():
        try:
            c.drawImage(str(settings.logo_path), 40, y - 30, 45, 35, mask="auto")
            c.setFont("Helvetica-Bold", 14)
            c.drawString(95, y, title)
            y -= 45
        except Exception:
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40, y, title)
            y -= 24
    else:
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, title)
        y -= 24
    c.setFont("Helvetica", 8)
    for i in items:
        line = (
            f"{i.numero} | {i.sender_name[:18]} | {i.recipient_name[:18]} | "
            f"{float(i.poids)}kg | {i.statut} | {float(i.total):.0f} FC"
        )
        if y < 40:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica", 8)
        c.drawString(40, y, line[:120])
        y -= 12
    c.save()
    return path


def export_financial_report_excel(
    financial_data: dict,
    expenses: list,
    path: Path,
    title: str = "Rapport Financier"
) -> Path:
    """Export complete financial report to Excel with professional accounting format."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    currency_alignment = Alignment(horizontal="right")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    
    # Title
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'] = "RAPPORT FINANCIER"
    ws_summary['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    
    ws_summary['A3'] = "Période:"
    ws_summary['B3'] = financial_data.get("period", "Non spécifié")
    ws_summary['A4'] = "Date d'export:"
    ws_summary['B4'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    ws_summary['A6'] = "Revenus totaux"
    ws_summary['B6'] = float(financial_data.get("revenue", {}).get("total", 0))
    ws_summary['B6'].number_format = '#,##0.00'
    ws_summary['B6'].font = total_font
    
    ws_summary['A7'] = "Ventes de billets"
    ws_summary['B7'] = float(financial_data.get("revenue", {}).get("tickets", {}).get("amount", 0))
    ws_summary['B7'].number_format = '#,##0.00'
    
    ws_summary['A8'] = "Ventes de bagages"
    ws_summary['B8'] = float(financial_data.get("revenue", {}).get("luggage", {}).get("amount", 0))
    ws_summary['B8'].number_format = '#,##0.00'
    
    ws_summary['A10'] = "Dépenses totales"
    ws_summary['B10'] = float(financial_data.get("expenses", 0))
    ws_summary['B10'].number_format = '#,##0.00'
    ws_summary['B10'].font = Font(bold=True, size=11, color="C00000")
    
    ws_summary['A11'] = "Bénéfice net"
    ws_summary['B11'] = float(financial_data.get("profit", 0))
    ws_summary['B11'].number_format = '#,##0.00'
    ws_summary['B11'].font = total_font
    if financial_data.get("profit", 0) >= 0:
        ws_summary['B11'].font = Font(bold=True, size=11, color="008000")
    else:
        ws_summary['B11'].font = Font(bold=True, size=11, color="C00000")
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # Expenses sheet
    ws_expenses = wb.create_sheet("Dépenses")
    headers = ["Date", "Catégorie", "Description", "Montant", "Mode de paiement", "Fournisseur", "Justificatif"]
    ws_expenses.append(headers)
    
    # Apply header style
    for col_num, header in enumerate(headers, 1):
        cell = ws_expenses.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add expenses data
    total_expenses = 0
    for row_num, exp in enumerate(expenses, 2):
        ws_expenses.append([
            exp.date_paiement.strftime("%d/%m/%Y") if exp.date_paiement else "",
            exp.categorie.replace("_", " ").title(),
            exp.description or "",
            float(exp.montant),
            exp.mode_paiement.replace("_", " ").title(),
            exp.fournisseur or "",
            exp.piece_jointe or "",
        ])
        
        total_expenses += float(exp.montant)
        
        # Apply currency format to amount column
        amount_cell = ws_expenses.cell(row=row_num, column=4)
        amount_cell.number_format = '#,##0.00'
        amount_cell.alignment = currency_alignment
    
    # Add total row
    total_row = len(expenses) + 2
    ws_expenses.cell(row=total_row, column=3, value="TOTAL DÉPENSES")
    ws_expenses.cell(row=total_row, column=3).font = total_font
    ws_expenses.cell(row=total_row, column=3).fill = total_fill
    ws_expenses.cell(row=total_row, column=4, value=total_expenses)
    ws_expenses.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws_expenses.cell(row=total_row, column=4).font = total_font
    ws_expenses.cell(row=total_row, column=4).fill = total_fill
    ws_expenses.cell(row=total_row, column=4).alignment = currency_alignment
    
    # Adjust column widths
    ws_expenses.column_dimensions['A'].width = 12
    ws_expenses.column_dimensions['B'].width = 18
    ws_expenses.column_dimensions['C'].width = 30
    ws_expenses.column_dimensions['D'].width = 15
    ws_expenses.column_dimensions['E'].width = 15
    ws_expenses.column_dimensions['F'].width = 20
    ws_expenses.column_dimensions['G'].width = 25
    
    # Category breakdown sheet
    ws_categories = wb.create_sheet("Par Catégorie")
    ws_categories.append(["Catégorie", "Nombre", "Total", "% du total"])
    
    # Apply header style
    for col_num in range(1, 5):
        cell = ws_categories.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Calculate category totals
    category_totals = {}
    category_counts = {}
    for exp in expenses:
        category_totals[exp.categorie] = category_totals.get(exp.categorie, 0) + float(exp.montant)
        category_counts[exp.categorie] = category_counts.get(exp.categorie, 0) + 1
    
    # Add category data
    for row_num, (category, total) in enumerate(sorted(category_totals.items(), key=lambda x: x[1], reverse=True), 2):
        count = category_counts[category]
        percentage = (total / total_expenses * 100) if total_expenses > 0 else 0
        
        ws_categories.append([
            category.replace("_", " ").title(),
            count,
            total,
            f"{percentage:.1f}%"
        ])
        
        # Apply currency format
        ws_categories.cell(row=row_num, column=3).number_format = '#,##0.00'
        ws_categories.cell(row=row_num, column=3).alignment = currency_alignment
        ws_categories.cell(row=row_num, column=4).alignment = Alignment(horizontal="right")
    
    # Add total row
    total_row = len(category_totals) + 2
    ws_categories.cell(row=total_row, column=1, value="TOTAL")
    ws_categories.cell(row=total_row, column=1).font = total_font
    ws_categories.cell(row=total_row, column=1).fill = total_fill
    ws_categories.cell(row=total_row, column=2, value=len(expenses))
    ws_categories.cell(row=total_row, column=2).font = total_font
    ws_categories.cell(row=total_row, column=2).fill = total_fill
    ws_categories.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    ws_categories.cell(row=total_row, column=3, value=total_expenses)
    ws_categories.cell(row=total_row, column=3).number_format = '#,##0.00'
    ws_categories.cell(row=total_row, column=3).font = total_font
    ws_categories.cell(row=total_row, column=3).fill = total_fill
    ws_categories.cell(row=total_row, column=3).alignment = currency_alignment
    ws_categories.cell(row=total_row, column=4, value="100.0%")
    ws_categories.cell(row=total_row, column=4).font = total_font
    ws_categories.cell(row=total_row, column=4).fill = total_fill
    ws_categories.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    
    # Adjust column widths
    ws_categories.column_dimensions['A'].width = 20
    ws_categories.column_dimensions['B'].width = 12
    ws_categories.column_dimensions['C'].width = 15
    ws_categories.column_dimensions['D'].width = 12
    
    wb.save(path)
    return path


def export_expenses_excel(expenses: list, path: Path) -> Path:
    """Export expenses list to Excel with professional accounting format."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dépenses"

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="8C6A00", end_color="8C6A00", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="F2EADF", end_color="F2EADF", fill_type="solid")
    currency_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Alternating row colours
    row_fill_odd  = PatternFill(start_color="FFFDF9", end_color="FFFDF9", fill_type="solid")
    row_fill_even = PatternFill(start_color="FFF6E8", end_color="FFF6E8", fill_type="solid")

    # ── Titre d'en-tête du rapport ────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value="RAPPORT DES DÉPENSES — NGOKAF TRANS")
    title_cell.font = Font(bold=True, size=15, color="8C6A00")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="F2EADF", end_color="F2EADF", fill_type="solid")
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:G2')
    date_cell = ws.cell(row=2, column=1,
        value=f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  —  {len(expenses)} dépense(s)")
    date_cell.font = Font(italic=True, size=10, color="6D6D6D")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Blank separator row
    ws.row_dimensions[3].height = 6

    # ── En-têtes des colonnes ─────────────────────────────────────────────────
    header_row = 4
    headers = ["Date", "Catégorie", "Description", "Montant (FC)", "Mode paiement", "Fournisseur", "Justificatif"]
    ws.append(headers)  # goes to row 4

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 26

    # ── Données des dépenses ──────────────────────────────────────────────────
    total_expenses = 0
    category_totals = {}
    category_counts = {}

    for idx, exp in enumerate(expenses):
        row_num = header_row + 1 + idx
        category = exp.categorie.replace("_", " ").title()
        amount = float(exp.montant)

        ws.append([
            exp.date_paiement.strftime("%d/%m/%Y") if exp.date_paiement else "",
            category,
            exp.description or "",
            amount,
            exp.mode_paiement.replace("_", " ").title(),
            exp.fournisseur or "",
            exp.piece_jointe or "",
        ])

        total_expenses += amount
        category_totals[category] = category_totals.get(category, 0) + amount
        category_counts[category] = category_counts.get(category, 0) + 1

        fill = row_fill_even if idx % 2 == 1 else row_fill_odd
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_num].height = 20

        # Currency format for amount
        amount_cell = ws.cell(row=row_num, column=4)
        amount_cell.number_format = '#,##0.00'
        amount_cell.alignment = currency_alignment

    # ── Ligne TOTAL GÉNÉRAL ───────────────────────────────────────────────────
    total_row = header_row + 1 + len(expenses)
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.border = thin_border
    ws.cell(row=total_row, column=1, value="TOTAL GÉNÉRAL").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=total_row, column=4, value=total_expenses).font = Font(bold=True, size=12, color="8C6A00")
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).alignment = currency_alignment
    ws.row_dimensions[total_row].height = 24

    # ── Section sous-totaux par catégorie ─────────────────────────────────────
    subtotal_start = total_row + 2

    ws.merge_cells(f'A{subtotal_start}:G{subtotal_start}')
    section_cell = ws.cell(row=subtotal_start, column=1)
    section_cell.value = "SOUS-TOTAUX PAR CATÉGORIE"
    section_cell.font = Font(bold=True, size=12, color="8C6A00")
    section_cell.alignment = Alignment(horizontal="center", vertical="center")
    section_cell.fill = PatternFill(start_color="F2EADF", end_color="F2EADF", fill_type="solid")
    ws.row_dimensions[subtotal_start].height = 26

    sub_header_row = subtotal_start + 1
    sub_headers = ["Catégorie", "Nb. dépenses", "", "Montant total (FC)", "Part (%)"]
    for col_idx, hdr in enumerate(sub_headers, 1):
        cell = ws.cell(row=sub_header_row, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[sub_header_row].height = 24

    for idx, (category, total) in enumerate(
        sorted(category_totals.items(), key=lambda x: x[1], reverse=True)
    ):
        row_num = sub_header_row + 1 + idx
        count = category_counts[category]
        percentage = (total / total_expenses * 100) if total_expenses > 0 else 0
        fill = row_fill_even if idx % 2 == 1 else row_fill_odd

        c1 = ws.cell(row=row_num, column=1, value=category)
        c1.font = Font(bold=True, size=10)
        c1.fill = fill
        c1.border = thin_border
        c1.alignment = Alignment(vertical="center")

        c2 = ws.cell(row=row_num, column=2, value=count)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.fill = fill
        c2.border = thin_border

        c4 = ws.cell(row=row_num, column=4, value=total)
        c4.number_format = '#,##0.00'
        c4.alignment = currency_alignment
        c4.font = Font(bold=True, size=10, color="DC3545")
        c4.fill = fill
        c4.border = thin_border

        c5 = ws.cell(row=row_num, column=5, value=f"{percentage:.1f}%")
        c5.alignment = Alignment(horizontal="right", vertical="center")
        c5.fill = fill
        c5.border = thin_border

        ws.row_dimensions[row_num].height = 20

    # ── Largeurs des colonnes ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 34
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 30

    # Figer à partir de la ligne 5 (après les 3 lignes d'en-tête)
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(path)
    return path


def export_revenue_expense_comparison(
    daily_data: list,
    path: Path,
    title: str = "Comparaison Revenus Dépenses"
) -> Path:
    """Export daily revenue vs expense comparison to Excel with professional format."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparaison"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    currency_alignment = Alignment(horizontal="right")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "COMPARAISON REVENUS VS DÉPENSES"
    ws['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ["Date", "Revenus", "Dépenses", "Bénéfice"]
    ws.append(headers)
    ws.append(headers)  # Add headers at row 3
    
    # Apply header style
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    total_revenue = 0
    total_expenses = 0
    total_profit = 0
    
    for row_num, data in enumerate(daily_data, 4):
        date_str = data["date"].strftime("%d/%m/%Y") if hasattr(data["date"], "strftime") else str(data["date"])
        revenue = float(data["revenue"])
        expenses = float(data["expenses"])
        profit = float(data["profit"])
        
        ws.append([date_str, revenue, expenses, profit])
        
        total_revenue += revenue
        total_expenses += expenses
        total_profit += profit
        
        # Apply currency format
        for col in [2, 3, 4]:
            cell = ws.cell(row=row_num, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = currency_alignment
    
    # Add total row
    total_row = len(daily_data) + 4
    ws.cell(row=total_row, column=1, value="TOTAL PÉRIODE")
    ws.cell(row=total_row, column=1).font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=2, value=total_revenue)
    ws.cell(row=total_row, column=2).number_format = '#,##0.00'
    ws.cell(row=total_row, column=2).font = total_font
    ws.cell(row=total_row, column=2).fill = total_fill
    ws.cell(row=total_row, column=2).alignment = currency_alignment
    ws.cell(row=total_row, column=3, value=total_expenses)
    ws.cell(row=total_row, column=3).number_format = '#,##0.00'
    ws.cell(row=total_row, column=3).font = total_font
    ws.cell(row=total_row, column=3).fill = total_fill
    ws.cell(row=total_row, column=3).alignment = currency_alignment
    ws.cell(row=total_row, column=4, value=total_profit)
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row, column=4).font = total_font
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).alignment = currency_alignment
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    wb.save(path)
    return path


def export_complete_financial_report(
    financial_data: dict,
    expenses: list,
    path,
    title: str = "Rapport Financier Complet",
) -> Path:
    """
    Export a complete financial report to a single Excel sheet:
      1. Revenue details + total
      2. Expense details (date, category, description, amount, payment mode, supplier, justificatif) + total
      3. Grand total (revenues - expenses = net profit)

    The period is taken from financial_data['period'].
    ``path`` can be a str or Path.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Financier"

    # ── Styles ────────────────────────────────────────────────────────────────
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    FILL_TITLE        = _fill("F2EADF")
    FILL_SECTION_REV  = _fill("8C6A00")   # doré — revenus
    FILL_SECTION_EXP  = _fill("DC3545")   # rouge — dépenses
    FILL_SECTION_TOT  = _fill("2E2E2E")   # sombre — total général
    FILL_ROW_ODD      = _fill("FFFDF9")
    FILL_ROW_EVEN     = _fill("FFF6E8")
    FILL_SUBTOTAL     = _fill("F2EADF")

    FONT_WHITE_BIG  = Font(bold=True, size=13, color="FFFFFF")
    FONT_NORMAL     = Font(size=11)
    FONT_BOLD       = Font(bold=True, size=11)

    NUM_FMT = "#,##0.00"
    COLS    = 7   # A..G

    def _section_header(row_num: int, text: str, fill: PatternFill, font: Font) -> int:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=COLS)
        c = ws.cell(row=row_num, column=1, value=text)
        c.font = font
        c.fill = fill
        c.alignment = center
        c.border = thin
        ws.row_dimensions[row_num].height = 24
        return row_num + 1

    def _column_headers(row_num: int, headers: list[str], fill: PatternFill) -> int:
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = fill
            c.alignment = center
            c.border = thin
        ws.row_dimensions[row_num].height = 22
        return row_num + 1

    def _subtotal_row(row_num: int, label: str, amount, fill: PatternFill,
                      label_col: int = 1, amount_col: int = 4,
                      font_label: Font = FONT_BOLD, font_amount: Font = FONT_BOLD) -> int:
        for col in range(1, COLS + 1):
            c = ws.cell(row=row_num, column=col)
            c.fill = fill
            c.border = thin
        lc = ws.cell(row=row_num, column=label_col, value=label)
        lc.font = font_label
        lc.fill = fill
        lc.alignment = left
        ac = ws.cell(row=row_num, column=amount_col, value=float(amount))
        ac.number_format = NUM_FMT
        ac.alignment = right
        ac.font = font_amount
        ac.fill = fill
        ws.row_dimensions[row_num].height = 22
        return row_num + 1

    def _blank(row_num: int, n: int = 1) -> int:
        for i in range(n):
            ws.row_dimensions[row_num + i].height = 10
        return row_num + n

    # ══════════════════════════════════════════════════════════════════════════
    # RAPPORT HEADER
    # ══════════════════════════════════════════════════════════════════════════
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    tc = ws.cell(row=r, column=1, value="RAPPORT FINANCIER COMPLET — NGOKAF TRANS")
    tc.font = Font(bold=True, size=16, color="8C6A00")
    tc.fill = FILL_TITLE
    tc.alignment = center
    ws.row_dimensions[r].height = 34
    r += 1

    period_label = financial_data.get("period", "")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    sc = ws.cell(row=r, column=1,
        value=f"Période : {period_label}   |   Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    sc.font = Font(italic=True, size=10, color="6D6D6D")
    sc.fill = FILL_TITLE
    sc.alignment = center
    ws.row_dimensions[r].height = 18
    r += 1
    r = _blank(r, 1)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1 : REVENUS
    # ══════════════════════════════════════════════════════════════════════════
    r = _section_header(r, "1.  DÉTAIL DES REVENUS", FILL_SECTION_REV, FONT_WHITE_BIG)
    r = _column_headers(r, ["Source", "Quantité", "Montant (FC)", "", "", "", ""], FILL_SECTION_REV)

    revenue = financial_data.get("revenue", {})
    rev_rows = [
        ("Ventes de billets", revenue.get("tickets", {}).get("count", 0),
                              float(revenue.get("tickets", {}).get("amount", 0))),
        ("Ventes de bagages", revenue.get("luggage", {}).get("count", 0),
                              float(revenue.get("luggage", {}).get("amount", 0))),
    ]
    for idx, (src, qty, amt) in enumerate(rev_rows):
        fill = FILL_ROW_EVEN if idx % 2 == 1 else FILL_ROW_ODD
        c1 = ws.cell(row=r, column=1, value=src)
        c1.fill = fill; c1.border = thin; c1.alignment = left; c1.font = FONT_NORMAL
        c2 = ws.cell(row=r, column=2, value=qty)
        c2.fill = fill; c2.border = thin; c2.alignment = center; c2.font = FONT_NORMAL
        c3 = ws.cell(row=r, column=3, value=amt)
        c3.fill = fill; c3.border = thin; c3.alignment = right
        c3.number_format = NUM_FMT
        c3.font = Font(size=11, color="8C6A00")
        for col in range(4, COLS + 1):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill; cc.border = thin
        ws.row_dimensions[r].height = 20
        r += 1

    total_rev = float(revenue.get("total", 0))
    r = _subtotal_row(r, "TOTAL REVENUS", total_rev, FILL_SUBTOTAL,
                      label_col=1, amount_col=3,
                      font_label=Font(bold=True, size=12, color="8C6A00"),
                      font_amount=Font(bold=True, size=12, color="8C6A00"))
    r = _blank(r, 1)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2 : DÉPENSES
    # ══════════════════════════════════════════════════════════════════════════
    r = _section_header(r, "2.  DÉTAIL DES DÉPENSES", FILL_SECTION_EXP, FONT_WHITE_BIG)
    exp_cols = ["Date", "Catégorie", "Description", "Montant (FC)", "Mode paiement", "Fournisseur", "Justificatif"]
    r = _column_headers(r, exp_cols, FILL_SECTION_EXP)

    total_expenses = 0.0
    for idx, exp in enumerate(expenses):
        amt = float(exp.montant)
        total_expenses += amt
        fill = FILL_ROW_EVEN if idx % 2 == 1 else FILL_ROW_ODD
        values = [
            exp.date_paiement.strftime("%d/%m/%Y") if exp.date_paiement else "",
            exp.categorie.replace("_", " ").title(),
            exp.description or "",
            amt,
            exp.mode_paiement.replace("_", " ").title(),
            exp.fournisseur or "",
            exp.piece_jointe or "",
        ]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = fill; c.border = thin
            c.alignment = right if col == 4 else (center if col == 1 else left)
            c.font = FONT_NORMAL
        ws.cell(row=r, column=4).number_format = NUM_FMT
        ws.cell(row=r, column=4).font = Font(size=11, color="DC3545")
        ws.row_dimensions[r].height = 20
        r += 1

    if not expenses:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
        nc = ws.cell(row=r, column=1, value="Aucune dépense pour cette période")
        nc.font = Font(italic=True, size=11, color="6D6D6D")
        nc.fill = FILL_ROW_ODD; nc.alignment = center; nc.border = thin
        ws.row_dimensions[r].height = 20
        r += 1

    r = _subtotal_row(r, "TOTAL DÉPENSES", total_expenses, FILL_SUBTOTAL,
                      label_col=1, amount_col=4,
                      font_label=Font(bold=True, size=12, color="DC3545"),
                      font_amount=Font(bold=True, size=12, color="DC3545"))
    r = _blank(r, 1)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3 : TOTAL GÉNÉRAL
    # ══════════════════════════════════════════════════════════════════════════
    r = _section_header(r, "3.  TOTAL GÉNÉRAL", FILL_SECTION_TOT, FONT_WHITE_BIG)

    profit = float(financial_data.get("profit", total_rev - total_expenses))
    grand_rows = [
        ("Total Revenus",  total_rev,       "8C6A00"),
        ("Total Dépenses", total_expenses,  "DC3545"),
        ("Bénéfice Net",   profit,          "28A745" if profit >= 0 else "DC3545"),
    ]
    for idx, (label, amount, color) in enumerate(grand_rows):
        fill = FILL_ROW_EVEN if idx % 2 == 1 else FILL_ROW_ODD
        is_last = (idx == len(grand_rows) - 1)
        if is_last:
            fill = _fill("F2EADF")
        for col in range(1, COLS + 1):
            c = ws.cell(row=r, column=col)
            c.fill = fill; c.border = thin
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True, size=12, color=color)
        c1.fill = fill; c1.alignment = left
        c2 = ws.cell(row=r, column=4, value=amount)
        c2.number_format = NUM_FMT
        c2.alignment = right
        c2.font = Font(bold=True, size=12 if not is_last else 14, color=color)
        c2.fill = fill
        ws.row_dimensions[r].height = 24 if not is_last else 28
        r += 1

    # ── Largeurs des colonnes ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 28

    # Freeze after row 3 (title + subtitle)
    ws.freeze_panes = "A4"

    wb.save(path)
    return path


def export_activity_report_excel(
    kpis: dict,
    routes: list,
    start_date,
    end_date,
    path,
    period_label: str = "",
) -> Path:
    """Export activity report to Excel with professional styling and colors."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport d'Activité"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    FILL_TITLE       = PatternFill(start_color="F2EADF", end_color="F2EADF", fill_type="solid")
    FILL_HEADER      = PatternFill(start_color="8C6A00", end_color="8C6A00", fill_type="solid")
    FILL_ROW_ODD     = PatternFill(start_color="FFFDF9", end_color="FFFDF9", fill_type="solid")
    FILL_ROW_EVEN    = PatternFill(start_color="FFF6E8", end_color="FFF6E8", fill_type="solid")
    FILL_TOTAL       = PatternFill(start_color="F2EADF", end_color="F2EADF", fill_type="solid")

    FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
    FONT_BOLD   = Font(bold=True, size=11)
    FONT_NORMAL = Font(size=11)
    NUM_FMT     = "#,##0.00"
    COLS        = 4

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    tc = ws.cell(row=r, column=1, value="RAPPORT D'ACTIVITÉ & PERFORMANCES — NGOKAF TRANS")
    tc.font = Font(bold=True, size=16, color="8C6A00")
    tc.fill = FILL_TITLE; tc.alignment = center
    ws.row_dimensions[r].height = 34
    r += 1

    p_str = f"Période : {period_label} ({start_date} au {end_date})" if period_label else f"Période : {start_date} au {end_date}"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    sc = ws.cell(row=r, column=1, value=f"{p_str}   |   Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    sc.font = Font(italic=True, size=10, color="6D6D6D")
    sc.fill = FILL_TITLE; sc.alignment = center
    ws.row_dimensions[r].height = 18
    r += 2

    # ── Section 1 : Indicateurs Clés (KPIs) ────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    sh1 = ws.cell(row=r, column=1, value="1. INDICATEURS CLÉS DE PERFORMANCE")
    sh1.font = Font(bold=True, size=13, color="FFFFFF")
    sh1.fill = FILL_HEADER; sh1.alignment = center; sh1.border = thin
    ws.row_dimensions[r].height = 24
    r += 1

    kpi_items = [
        ("Recettes Totales", float(kpis.get("recettes_total", 0)), True),
        ("Ventes Billets (Quantité)", kpis.get("nb_billets", 0), False),
        ("Recettes Billets", float(kpis.get("recettes_billets", 0)), True),
        ("Ventes Bagages (Quantité)", kpis.get("nb_bagages", 0), False),
        ("Recettes Bagages", float(kpis.get("recettes_bagages", 0)), True),
    ]

    for idx, (label, val, is_currency) in enumerate(kpi_items):
        fill = FILL_ROW_EVEN if idx % 2 == 1 else FILL_ROW_ODD
        for col in range(1, COLS + 1):
            c = ws.cell(row=r, column=col)
            c.fill = fill; c.border = thin
        ws.cell(row=r, column=1, value=label).font = FONT_BOLD if idx == 0 else FONT_NORMAL
        ws.cell(row=r, column=1).alignment = left
        vc = ws.cell(row=r, column=3, value=val)
        vc.alignment = right if is_currency else center
        vc.font = Font(bold=True, size=11, color="8C6A00" if is_currency else "2E2E2E")
        if is_currency:
            vc.number_format = NUM_FMT
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1

    # ── Section 2 : Performance par Trajet ────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    sh2 = ws.cell(row=r, column=1, value="2. PERFORMANCE DÉTAILLÉE PAR TRAJET")
    sh2.font = Font(bold=True, size=13, color="FFFFFF")
    sh2.fill = FILL_HEADER; sh2.alignment = center; sh2.border = thin
    ws.row_dimensions[r].height = 24
    r += 1

    headers = ["Trajet", "Billets Vendus", "Recettes Totales (FC)", "Part (%)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = center; c.border = thin
    ws.row_dimensions[r].height = 22
    r += 1

    total_revenue_routes = sum(float(row[2]) for row in routes) if routes else 0.0
    total_tickets_routes = sum(int(row[1]) for row in routes) if routes else 0

    for idx, (label, count, amount) in enumerate(routes):
        amt_val = float(amount)
        cnt_val = int(count)
        pct = (amt_val / total_revenue_routes * 100) if total_revenue_routes > 0 else 0.0
        fill = FILL_ROW_EVEN if idx % 2 == 1 else FILL_ROW_ODD

        ws.cell(row=r, column=1, value=label).alignment = left
        ws.cell(row=r, column=2, value=cnt_val).alignment = center
        c3 = ws.cell(row=r, column=3, value=amt_val)
        c3.alignment = right; c3.number_format = NUM_FMT; c3.font = Font(size=11, color="8C6A00")
        ws.cell(row=r, column=4, value=f"{pct:.1f}%").alignment = right

        for col in range(1, COLS + 1):
            c = ws.cell(row=r, column=col)
            c.fill = fill; c.border = thin
            if col != 3:
                c.font = FONT_NORMAL
        ws.row_dimensions[r].height = 20
        r += 1

    # Total row
    for col in range(1, COLS + 1):
        c = ws.cell(row=r, column=col)
        c.fill = FILL_TOTAL; c.border = thin
    ws.cell(row=r, column=1, value="TOTAL TRAJETS").font = FONT_BOLD
    ws.cell(row=r, column=2, value=total_tickets_routes).font = FONT_BOLD; ws.cell(row=r, column=2).alignment = center
    t3 = ws.cell(row=r, column=3, value=total_revenue_routes)
    t3.font = Font(bold=True, size=12, color="8C6A00"); t3.alignment = right; t3.number_format = NUM_FMT
    ws.cell(row=r, column=4, value="100.0%").font = FONT_BOLD; ws.cell(row=r, column=4).alignment = right
    ws.row_dimensions[r].height = 24

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 15

    ws.freeze_panes = "A4"
    wb.save(path)
    return path


def export_activity_report_pdf(
    kpis: dict,
    routes: list,
    start_date,
    end_date,
    path,
    period_label: str = "",
) -> Path:
    """Export activity report to a clean, highly stylized PDF using ReportLab Platypus."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    primary_color = colors.HexColor("#8C6A00")
    secondary_color = colors.HexColor("#2F2A24")
    bg_light = colors.HexColor("#FFFDF9")
    bg_even = colors.HexColor("#FFF6E8")
    border_color = colors.HexColor("#E4D8C3")

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=primary_color,
        alignment=1, # Center
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#6D6D6D"),
        alignment=1,
    )
    section_style = ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=16,
        textColor=primary_color,
        spaceBefore=10,
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=secondary_color,
    )
    cell_bold_style = ParagraphStyle(
        "TableCellBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=12,
        textColor=secondary_color,
    )
    cell_right_bold = ParagraphStyle(
        "CellRightBold",
        parent=cell_bold_style,
        alignment=2,
        textColor=primary_color,
    )

    story = []

    # Logo & Header Banner
    if settings.logo_path.exists():
        try:
            logo_img = RLImage(str(settings.logo_path), width=70, height=56)
            logo_img.hAlign = 'CENTER'
            story.append(logo_img)
            story.append(Spacer(1, 4))
        except Exception:
            pass

    # Title & Header
    story.append(Paragraph("NGOKAF TRANS", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph("RAPPORT D'ACTIVITÉ & PERFORMANCES", ParagraphStyle("SubTitle2", parent=title_style, fontSize=13, leading=16, textColor=secondary_color)))
    story.append(Spacer(1, 4))
    p_text = f"Période : {period_label} ({start_date} au {end_date})" if period_label else f"Période du {start_date} au {end_date}"
    story.append(Paragraph(f"{p_text} — Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1.5, color=primary_color, spaceBefore=0, spaceAfter=12))

    # Section 1: KPIs
    story.append(Paragraph("1. Indicateurs Clés de Performance", section_style))

    kpi_data = [
        [Paragraph("<b>Indicateur</b>", cell_style), Paragraph("<b>Valeur</b>", cell_style)],
        [Paragraph("Recettes Totales", cell_style), Paragraph(f"<b>{format_fc(kpis.get('recettes_total', 0))}</b>", cell_right_bold)],
        [Paragraph("Ventes Billets (Quantité)", cell_style), Paragraph(str(kpis.get("nb_billets", 0)), cell_style)],
        [Paragraph("Recettes Billets", cell_style), Paragraph(format_fc(kpis.get("recettes_billets", 0)), cell_style)],
        [Paragraph("Ventes Bagages (Quantité)", cell_style), Paragraph(str(kpis.get("nb_bagages", 0)), cell_style)],
        [Paragraph("Recettes Bagages", cell_style), Paragraph(format_fc(kpis.get("recettes_bagages", 0)), cell_style)],
    ]

    t1 = Table(kpi_data, colWidths=[300, 220])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [bg_light, bg_even]),
        ('GRID', (0, 0), (-1, -1), 0.5, border_color),
    ]))
    # Header text color fix
    kpi_data[0][0] = Paragraph("<font color='white'><b>Indicateur</b></font>", cell_style)
    kpi_data[0][1] = Paragraph("<font color='white'><b>Valeur</b></font>", cell_style)

    story.append(t1)
    story.append(Spacer(1, 16))

    # Section 2: Performance par trajet
    story.append(Paragraph("2. Performance par Trajet", section_style))

    route_table_data = [
        [
            Paragraph("<font color='white'><b>Trajet</b></font>", cell_style),
            Paragraph("<font color='white'><b>Billets</b></font>", cell_style),
            Paragraph("<font color='white'><b>Recettes Totales</b></font>", cell_style),
        ]
    ]

    tot_tickets = 0
    tot_amount = 0.0

    for label, count, amount in routes:
        cnt_val = int(count)
        amt_val = float(amount)
        tot_tickets += cnt_val
        tot_amount += amt_val
        route_table_data.append([
            Paragraph(label, cell_style),
            Paragraph(str(cnt_val), cell_style),
            Paragraph(format_fc(amt_val), cell_bold_style),
        ])

    route_table_data.append([
        Paragraph("<b>TOTAL TRAJETS</b>", cell_bold_style),
        Paragraph(f"<b>{tot_tickets}</b>", cell_bold_style),
        Paragraph(f"<b>{format_fc(tot_amount)}</b>", cell_right_bold),
    ])

    doc.build(story)
    return path


def export_drivers_excel(drivers: list, path) -> Path:
    """Export drivers list to Excel with professional accounting format."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Conducteurs"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    FILL_TITLE    = PatternFill(start_color="F2EADF", end_color="F2EADF", fill_type="solid")
    FILL_HEADER   = PatternFill(start_color="8C6A00", end_color="8C6A00", fill_type="solid")
    FILL_ROW_ODD  = PatternFill(start_color="FFFDF9", end_color="FFFDF9", fill_type="solid")
    FILL_ROW_EVEN = PatternFill(start_color="FFF6E8", end_color="FFF6E8", fill_type="solid")

    FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
    FONT_NORMAL = Font(size=11)
    COLS        = 8

    # Title
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    tc = ws.cell(row=r, column=1, value="REGISTRE DES CONDUCTEURS — NGOKAF TRANS")
    tc.font = Font(bold=True, size=16, color="8C6A00")
    tc.fill = FILL_TITLE; tc.alignment = center
    ws.row_dimensions[r].height = 34
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    sc = ws.cell(row=r, column=1, value=f"Total : {len(drivers)} conducteur(s)   |   Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    sc.font = Font(italic=True, size=10, color="6D6D6D")
    sc.fill = FILL_TITLE; sc.alignment = center
    ws.row_dimensions[r].height = 18
    r += 2

    # Headers
    headers = ["ID", "Nom & Prénom", "Téléphone", "N° Permis", "Date Exp. Permis", "Bus Assigné", "Disponibilité", "Statut"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = center; c.border = thin
    ws.row_dimensions[r].height = 24
    r += 1

    header_row = r - 1

    for idx, d in enumerate(drivers):
        fill = FILL_ROW_EVEN if idx % 2 == 1 else FILL_ROW_ODD
        exp_date = d.date_expiration_permis.strftime("%d/%m/%Y") if getattr(d, "date_expiration_permis", None) else "—"
        bus_code = d.bus.code if getattr(d, "bus", None) else "—"
        tel = d.telephone or "—"
        permis = d.numero_permis or "—"
        dispo = d.disponibilite.replace("_", " ").title() if getattr(d, "disponibilite", None) else "Dispo"
        statut = d.statut.title() if getattr(d, "statut", None) else "Actif"

        row_vals = [d.id, d.full_name, tel, permis, exp_date, bus_code, dispo, statut]

        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = fill; c.border = thin; c.font = FONT_NORMAL
            if col in [1, 3, 4, 5, 6, 7, 8]:
                c.alignment = center
            else:
                c.alignment = left

        ws.row_dimensions[r].height = 20
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12

    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)
    return path


def export_drivers_pdf(drivers: list, path) -> Path:
    """Export drivers list to PDF with styled profile cards and avatar photos."""
    import os
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    primary_color   = colors.HexColor("#8C6A00")
    secondary_color = colors.HexColor("#2F2A24")
    card_bg         = colors.HexColor("#FFFDF9")
    card_border     = colors.HexColor("#E4D8C3")

    title_style = ParagraphStyle(
        "CardTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=primary_color,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        "CardSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#6D6D6D"),
        alignment=1,
    )
    driver_name_style = ParagraphStyle(
        "DriverName",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=14,
        textColor=primary_color,
    )
    info_style = ParagraphStyle(
        "DriverInfo",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=12,
        textColor=secondary_color,
    )

    story = []

    # Logo Banner
    if settings.logo_path.exists():
        try:
            logo_img = RLImage(str(settings.logo_path), width=70, height=56)
            logo_img.hAlign = 'CENTER'
            story.append(logo_img)
            story.append(Spacer(1, 4))
        except Exception:
            pass

    # Title Banner
    story.append(Paragraph("NGOKAF TRANS", title_style))
    story.append(Spacer(1, 3))
    story.append(Paragraph("REGISTRE & CARTES DES CONDUCTEURS", ParagraphStyle("Sub2", parent=title_style, fontSize=13, textColor=secondary_color)))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Total : {len(drivers)} conducteur(s) — Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=1.5, color=primary_color, spaceBefore=0, spaceAfter=14))

    # Helper function to generate a single driver card flowable
    def _create_driver_card(d) -> Table:
        # Check photo
        photo_flowable = None
        if getattr(d, "photo_path", None) and os.path.exists(d.photo_path):
            try:
                photo_flowable = RLImage(d.photo_path, width=54, height=54)
            except Exception:
                photo_flowable = None

        if not photo_flowable:
            # Create text avatar
            initials = "".join([p[0] for p in d.full_name.split()[:2]]).upper() or "?"
            photo_flowable = Paragraph(f"<font size=16 color='#8C6A00'><b>{initials}</b></font>", ParagraphStyle("Init", alignment=1))

        # Details
        tel = d.telephone or "—"
        bus_code = d.bus.code if getattr(d, "bus", None) else "Aucun"
        permis = d.numero_permis or "—"
        dispo = d.disponibilite.replace("_", " ").title()
        statut = d.statut.title()

        dispo_color = "#10B981" if d.disponibilite == "disponible" else ("#F97316" if d.disponibilite == "en_route" else "#EF4444")
        statut_color = "#28A745" if d.statut == "actif" else "#6C757D"

        details_html = (
            f"<b>Tél :</b> {tel}<br/>"
            f"<b>Bus :</b> {bus_code}<br/>"
            f"<b>Permis :</b> {permis}<br/>"
            f"<b>Dispo :</b> <font color='{dispo_color}'><b>{dispo}</b></font>  |  "
            f"<b>Statut :</b> <font color='{statut_color}'><b>{statut}</b></font>"
        )

        card_content = [
            [Paragraph(f"<b>{d.full_name}</b>", driver_name_style), ""],
            [photo_flowable, Paragraph(details_html, info_style)]
        ]

        t_card = Table(card_content, colWidths=[60, 185])
        t_card.setStyle(TableStyle([
            ('SPAN', (0, 0), (1, 0)),
            ('BACKGROUND', (0, 0), (-1, -1), card_bg),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 1), (0, 1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, card_border),
            ('LINEBELOW', (0, 0), (1, 0), 0.5, colors.HexColor("#F2EADF")),
        ]))
        return t_card

    # Create grid of cards (2 cards per row)
    cards_grid = []
    current_row = []

    for idx, d in enumerate(drivers):
        card = _create_driver_card(d)
        current_row.append(card)

        if len(current_row) == 2:
            cards_grid.append(current_row)
            current_row = []

    if current_row:
        if len(current_row) == 1:
            current_row.append("")  # Empty cell placeholder
        cards_grid.append(current_row)

    if cards_grid:
        grid_table = Table(cards_grid, colWidths=[260, 260])
        grid_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(grid_table)
    else:
        story.append(Paragraph("Aucun conducteur trouvé.", subtitle_style))

    doc.build(story)
    return path

def export_users_excel(users: list, path) -> Path:
    """Export users list to Excel with professional format."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    FILL_TITLE    = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    FILL_HEADER   = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    FILL_ROW_ODD  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    FILL_ROW_EVEN = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    FILL_ACTIF    = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    FILL_BLOQUE   = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    FILL_ADMIN    = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
    FONT_NORMAL = Font(size=11)
    COLS        = 8

    # Title
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    tc = ws.cell(row=r, column=1, value="REGISTRE DES UTILISATEURS — NGOKAF TRANS")
    tc.font = Font(bold=True, size=16, color="1A73E8")
    tc.fill = FILL_TITLE; tc.alignment = center
    ws.row_dimensions[r].height = 34
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    sc = ws.cell(row=r, column=1, value=f"Total : {len(users)} utilisateur(s)   |   Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    sc.font = Font(italic=True, size=10, color="6D6D6D")
    sc.fill = FILL_TITLE; sc.alignment = center
    ws.row_dimensions[r].height = 18
    r += 2

    # Headers
    headers = ["ID", "Nom & Prénom", "Identifiant", "Téléphone", "Email", "Adresse", "Rôle", "Statut"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = center; c.border = thin
    ws.row_dimensions[r].height = 24
    header_row = r
    r += 1

    for idx, u in enumerate(users):
        is_admin = getattr(u, "role", "") == "administrateur"
        statut = getattr(u, "statut", "actif")
        is_bloque = statut == "bloque"

        if is_admin:
            fill = FILL_ADMIN
        elif is_bloque:
            fill = FILL_BLOQUE
        else:
            fill = FILL_ROW_EVEN if idx % 2 == 1 else FILL_ROW_ODD

        last_login = u.last_login.strftime("%d/%m/%Y %H:%M") if getattr(u, "last_login", None) else "—"
        tel = u.telephone or "—"
        email = getattr(u, "email", None) or "—"
        adresse = getattr(u, "adresse", None) or "—"
        role = u.role.title()
        statut_label = "Bloqué" if is_bloque else "Actif"

        row_vals = [u.id, u.full_name, u.username, tel, email, adresse, role, statut_label]

        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = fill; c.border = thin; c.font = FONT_NORMAL
            if col in [1, 3, 7, 8]:
                c.alignment = center
            else:
                c.alignment = left

        # Color code statut cell
        statut_cell = ws.cell(row=r, column=8)
        if is_bloque:
            statut_cell.font = Font(size=11, bold=True, color="C5221F")
        else:
            statut_cell.font = Font(size=11, bold=True, color="137333")

        ws.row_dimensions[r].height = 20
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12

    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)
    return path


def export_users_pdf(users: list, path) -> Path:
    """Export users list to PDF with styled profile cards."""
    import os
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    primary_color   = colors.HexColor("#1A73E8")
    secondary_color = colors.HexColor("#202124")
    card_bg         = colors.HexColor("#F8FAFF")
    card_border     = colors.HexColor("#C5D5F5")
    admin_bg        = colors.HexColor("#FFF8E7")
    admin_border    = colors.HexColor("#FBBC04")

    title_style = ParagraphStyle(
        "UserTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=primary_color,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        "UserSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#6D6D6D"),
        alignment=1,
    )
    user_name_style = ParagraphStyle(
        "UserName",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=14,
        textColor=primary_color,
    )
    info_style = ParagraphStyle(
        "UserInfo",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=12,
        textColor=secondary_color,
    )

    story = []

    # Logo Banner
    if settings.logo_path.exists():
        try:
            logo_img = RLImage(str(settings.logo_path), width=70, height=56)
            logo_img.hAlign = 'CENTER'
            story.append(logo_img)
            story.append(Spacer(1, 4))
        except Exception:
            pass

    # Title Banner
    story.append(Paragraph("NGOKAF TRANS", title_style))
    story.append(Spacer(1, 3))
    story.append(Paragraph("REGISTRE DES UTILISATEURS &amp; CAISSIERS", ParagraphStyle("Sub2", parent=title_style, fontSize=13, textColor=secondary_color)))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Total : {len(users)} utilisateur(s) — Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=1.5, color=primary_color, spaceBefore=0, spaceAfter=14))

    def _create_user_card(u) -> Table:
        # Photo
        photo_flowable = None
        if getattr(u, "photo_path", None) and os.path.exists(u.photo_path):
            try:
                photo_flowable = RLImage(u.photo_path, width=54, height=54)
            except Exception:
                photo_flowable = None

        if not photo_flowable:
            initials = "".join([p[0] for p in u.full_name.split()[:2]]).upper() or "?"
            init_color = "#1A73E8" if getattr(u, "role", "") == "caissier" else "#FBBC04"
            photo_flowable = Paragraph(
                f"<font size=16 color='{init_color}'><b>{initials}</b></font>",
                ParagraphStyle("Init", alignment=1)
            )

        is_admin = getattr(u, "role", "") == "administrateur"
        statut = getattr(u, "statut", "actif")
        tel = u.telephone or "—"
        email = getattr(u, "email", None) or "—"
        adresse = getattr(u, "adresse", None) or "—"
        role_label = "Administrateur" if is_admin else "Caissier"
        statut_label = "Bloqué" if statut == "bloque" else "Actif"
        last_login = u.last_login.strftime("%d/%m/%Y %H:%M") if getattr(u, "last_login", None) else "Jamais"

        statut_color  = "#EF4444" if statut == "bloque" else "#10B981"
        role_color    = "#F59E0B" if is_admin else "#3B82F6"

        details_html = (
            f"<b>Login :</b> {u.username}<br/>"
            f"<b>Tél :</b> {tel}  |  <b>Email :</b> {email}<br/>"
            f"<b>Adresse :</b> {adresse}<br/>"
            f"<b>Rôle :</b> <font color='{role_color}'><b>{role_label}</b></font>  |  "
            f"<b>Statut :</b> <font color='{statut_color}'><b>{statut_label}</b></font><br/>"
            f"<b>Dernière connexion :</b> {last_login}"
        )

        bg_color   = admin_bg   if is_admin else card_bg
        bdr_color  = admin_border if is_admin else card_border

        card_content = [
            [Paragraph(f"<b>{u.full_name}</b>", user_name_style), ""],
            [photo_flowable, Paragraph(details_html, info_style)]
        ]

        t_card = Table(card_content, colWidths=[65, 180])
        t_card.setStyle(TableStyle([
            ('SPAN', (0, 0), (1, 0)),
            ('BACKGROUND', (0, 0), (-1, -1), bg_color),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 1), (0, 1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, bdr_color),
            ('LINEBELOW', (0, 0), (1, 0), 0.5, colors.HexColor("#E8F0FE")),
        ]))
        return t_card

    # Grid of cards (2 per row)
    cards_grid = []
    current_row = []

    for u in users:
        card = _create_user_card(u)
        current_row.append(card)
        if len(current_row) == 2:
            cards_grid.append(current_row)
            current_row = []

    if current_row:
        if len(current_row) == 1:
            current_row.append("")
        cards_grid.append(current_row)

    if cards_grid:
        grid_table = Table(cards_grid, colWidths=[260, 260])
        grid_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(grid_table)
    else:
        story.append(Paragraph("Aucun utilisateur trouvé.", subtitle_style))

    doc.build(story)
    return path
